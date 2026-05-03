import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\vansh\Desktop\MAIN\homeguru\Study Area Categories (1).xlsx', data_only=True)
with open('excel_dump.txt', 'w', encoding='utf-8') as f:
    for name in wb.sheetnames:
        ws = wb[name]
        f.write(f'=== SHEET: {name} ===\n')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if any(c is not None for c in row):
                f.write(str(list(row)) + '\n')
        f.write('\n')
print('Done')
